"""
Standalone RAG evaluation runner — produces JSON, Excel, and HTML reports.

Runs every question from testset.csv against the live /chat/inspect endpoint,
evaluates with 4 DeepEval RAG metrics, and saves results to:
  - testing/reports/rag_eval_report_YYYY-MM-DD.json
  - testing/reports/rag_eval_report_YYYY-MM-DD.html  ← open in any browser
  - testing/Testing.xlsx  ("RAG Eval" sheet — created if missing, replaced if exists)

Prerequisites:
  1. Backend running:  uvicorn server:app --port 8001  (from sg-property-agent/backend/)
  2. OPENAI_API_KEY set in testing/evaluation/.env.eval

Usage:
    cd "c:\\GEEMETH\\N\\Property Advisory AI Agent"
    python testing/evaluation/run_rag_eval.py
"""

import csv
import html as html_escape_module
import json
import os
import time
import uuid
from datetime import datetime, date
from pathlib import Path

import requests
from deepeval.metrics import (
    AnswerRelevancyMetric,
    ContextualPrecisionMetric,
    ContextualRecallMetric,
    FaithfulnessMetric,
)
from deepeval.test_case import LLMTestCase
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env.eval", override=True)

# ── Config ────────────────────────────────────────────────────────────────────

_BACKEND_URL  = os.getenv("BACKEND_URL", "http://localhost:8001")
_KB_URL       = os.getenv("KB_PIPELINE_URL", "http://localhost:8000")
_TESTSET_FILE = Path(__file__).parent / "testset.csv"
_REPORTS_DIR  = Path(__file__).resolve().parents[2] / "testing" / "reports"
_EXCEL_FILE   = Path(__file__).resolve().parents[2] / "testing" / "Testing.xlsx"
_SHEET_NAME   = "RAG Eval"
_TIMEOUT_S    = 120
_THRESHOLD    = 0.8

_METRIC_KEYS = ["Faithfulness", "AnswerRelevancy", "ContextualPrecision", "ContextualRecall"]

# ── Load test set ─────────────────────────────────────────────────────────────

def _load_testset() -> list[dict]:
    if not _TESTSET_FILE.exists():
        raise FileNotFoundError(
            f"testset.csv not found at {_TESTSET_FILE}."
        )
    records = []
    with _TESTSET_FILE.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            records.append({
                "Source":          row.get("Source", ""),
                "input":           row.get("Question", ""),
                "expected_output": row.get("Answer", ""),
                "context":         [],
            })
    if not records:
        raise ValueError("testset.csv is empty.")
    return records


# ── API call ──────────────────────────────────────────────────────────────────

def _call_inspect(question: str) -> dict:
    payload = {"question": question, "thread_id": str(uuid.uuid4())}
    resp = requests.post(
        f"{_BACKEND_URL}/chat/inspect",
        json=payload,
        timeout=_TIMEOUT_S,
    )
    resp.raise_for_status()
    return resp.json()


def _extract_retrieval_context(chunks_by_agent: dict) -> list[str]:
    texts = []
    for agent_chunks in chunks_by_agent.values():
        for chunk in agent_chunks:
            text = chunk.get("text", "").strip()
            if text:
                texts.append(text)
    return texts


def _check_kb_environment() -> str:
    """Call KB-Pipeline /config and return the active Pinecone index name."""
    try:
        resp = requests.get(f"{_KB_URL}/config", timeout=5)
        resp.raise_for_status()
        index = resp.json().get("pinecone_index", "unknown")
        expected = os.getenv("PINECONE_INDEX", "")
        status = ""
        if expected and index != expected:
            status = f"  WARNING: expected '{expected}'"
        print(f"  KB-Pipeline      : {_KB_URL}")
        print(f"  Pinecone index   : {index}{status}")
        return index
    except Exception as exc:
        print(f"  KB-Pipeline      : {_KB_URL}  (health check failed: {exc})")
        return "unknown"


# ── Evaluate one case ─────────────────────────────────────────────────────────

def _evaluate_case(test_id: str, rec: dict) -> dict:
    question        = rec["input"]
    expected_output = rec["expected_output"]
    synth_context   = [str(c) for c in rec["context"]] if rec["context"] else []

    print(f"\n{'─'*60}")
    print(f"[{test_id}] {question}")

    t0 = time.time()
    try:
        result = _call_inspect(question)
    except Exception as exc:
        print(f"  API ERROR: {exc}")
        return {
            "id": test_id, "question": question,
            "actual_output": "", "expected_output": expected_output,
            "retrieved_chunks": [], "metric_reasons": {},
            "error": str(exc), "latency_s": round(time.time() - t0, 2),
            "overall_passed": False,
            **{k: {"score": 0.0, "passed": False} for k in _METRIC_KEYS},
        }

    actual_output     = result.get("answer", "")
    retrieval_context = _extract_retrieval_context(result.get("retrieved_chunks", {}))
    if not retrieval_context:
        retrieval_context = synth_context

    print(f"  Expected : {expected_output[:120]}{'…' if len(expected_output) > 120 else ''}")
    print(f"  Answer   : {actual_output[:120]}{'…' if len(actual_output) > 120 else ''}")
    print(f"  Chunks   : {len(retrieval_context)} retrieved")
    for i, chunk in enumerate(retrieval_context[:3], 1):
        print(f"    [{i}] {chunk[:150]}{'…' if len(chunk) > 150 else ''}")
    if len(retrieval_context) > 3:
        print(f"    … and {len(retrieval_context) - 3} more")

    test_case = LLMTestCase(
        input=question,
        actual_output=actual_output,
        expected_output=expected_output,
        retrieval_context=retrieval_context,
    )

    metrics = [
        FaithfulnessMetric(threshold=_THRESHOLD),
        AnswerRelevancyMetric(threshold=_THRESHOLD),
        ContextualPrecisionMetric(threshold=_THRESHOLD),
        ContextualRecallMetric(threshold=_THRESHOLD),
    ]

    metric_results = {}
    metric_reasons = {}
    for metric in metrics:
        key = metric.__class__.__name__
        try:
            metric.measure(test_case)
            score  = round(metric.score, 4)
            passed = metric.is_successful()
            reason = getattr(metric, "reason", "") or ""
            metric_results[key] = {"score": score, "passed": passed}
            metric_reasons[key] = reason
            status_icon = "✓" if passed else "✗"
            print(f"  {key:<30} {score:.2f} {status_icon}")
            if reason:
                print(f"    → {reason[:200]}{'…' if len(reason) > 200 else ''}")
        except Exception as exc:
            metric_results[key] = {"score": 0.0, "passed": False, "error": str(exc)}
            metric_reasons[key] = f"ERROR: {exc}"
            print(f"  {key:<30} ERROR: {exc}")

    overall_passed = all(v["passed"] for v in metric_results.values())
    latency_s      = round(time.time() - t0, 2)

    status = "PASS ✓" if overall_passed else "FAIL ✗"
    print(f"  ── {status}  ({latency_s}s)")

    return {
        "id":               test_id,
        "question":         question,
        "actual_output":    actual_output,
        "expected_output":  expected_output,
        "retrieved_chunks": retrieval_context,
        "metric_reasons":   metric_reasons,
        "latency_s":        latency_s,
        "overall_passed":   overall_passed,
        **metric_results,
    }


# ── JSON report ───────────────────────────────────────────────────────────────

def _save_json(cases: list[dict]) -> Path:
    _REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    run_date    = date.today().isoformat()
    report_path = _REPORTS_DIR / f"rag_eval_report_{run_date}.json"

    total  = len(cases)
    passed = sum(1 for c in cases if c.get("overall_passed"))

    metric_class_keys = [
        "FaithfulnessMetric", "AnswerRelevancyMetric",
        "ContextualPrecisionMetric", "ContextualRecallMetric",
    ]
    # Support both plain-name and class-name keys
    def _get_metric(c, key):
        return c.get(key) or c.get(key.replace("Metric", "")) or {}

    metric_agg = {}
    for key in _METRIC_KEYS:
        class_key = key + "Metric"
        scores  = []
        passing = []
        for c in cases:
            m = c.get(class_key) or c.get(key) or {}
            if m:
                scores.append(m.get("score", 0.0))
                passing.append(m.get("passed", False))
        metric_agg[key] = {
            "mean":      round(sum(scores) / len(scores), 4) if scores else 0.0,
            "pass_rate": round(sum(passing) / len(passing), 4) if passing else 0.0,
        }

    report = {
        "generated_at":  datetime.now().isoformat(),
        "backend":       _BACKEND_URL,
        "total":         total,
        "passed":        passed,
        "pass_rate":     round(passed / total, 4) if total else 0.0,
        "metric_scores": metric_agg,
        "cases": [
            {
                "id":               c["id"],
                "question":         c["question"],
                "actual_output":    c.get("actual_output", ""),
                "expected_output":  c.get("expected_output", ""),
                "retrieved_chunks": c.get("retrieved_chunks", []),
                "metric_reasons":   c.get("metric_reasons", {}),
                "faithfulness":         c.get("FaithfulnessMetric", {}),
                "answer_relevancy":     c.get("AnswerRelevancyMetric", {}),
                "contextual_precision": c.get("ContextualPrecisionMetric", {}),
                "contextual_recall":    c.get("ContextualRecallMetric", {}),
                "overall_passed":   c.get("overall_passed", False),
                "latency_s":        c.get("latency_s", 0.0),
            }
            for c in cases
        ],
    }

    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    return report_path


# ── HTML report ───────────────────────────────────────────────────────────────

def _save_html(cases: list[dict], run_date: str, active_index: str = "unknown") -> Path:
    _REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    report_path = _REPORTS_DIR / f"rag_eval_report_{run_date}.html"

    total  = len(cases)
    passed = sum(1 for c in cases if c.get("overall_passed"))
    pass_pct = round(passed / total * 100, 1) if total else 0.0

    def _metric(c, name):
        return c.get(f"{name}Metric") or c.get(name) or {}

    def _mean(key):
        scores = [_metric(c, key).get("score", 0.0) for c in cases]
        return round(sum(scores) / len(scores), 2) if scores else 0.0

    f_mean  = _mean("Faithfulness")
    ar_mean = _mean("AnswerRelevancy")
    cp_mean = _mean("ContextualPrecision")
    cr_mean = _mean("ContextualRecall")

    def score_color(score):
        return "#27ae60" if score >= _THRESHOLD else "#e74c3c"

    def h(text):
        return html_escape_module.escape(str(text))

    def _chunk_rows(chunks):
        if not chunks:
            return "<p style='color:#888;font-style:italic'>No chunks retrieved</p>"
        rows = []
        for i, chunk in enumerate(chunks, 1):
            rows.append(
                f"<div style='background:#f8f9fa;border-left:3px solid #3498db;"
                f"padding:8px 12px;margin:4px 0;font-size:12px;border-radius:2px'>"
                f"<strong>[{i}]</strong> {h(chunk[:500])}{'…' if len(chunk) > 500 else ''}</div>"
            )
        return "".join(rows)

    def _reason_rows(reasons):
        if not reasons:
            return ""
        metric_labels = {
            "FaithfulnessMetric": "Faithfulness",
            "AnswerRelevancyMetric": "Answer Relevancy",
            "ContextualPrecisionMetric": "Contextual Precision",
            "ContextualRecallMetric": "Contextual Recall",
        }
        rows = []
        for key, label in metric_labels.items():
            reason = reasons.get(key, "")
            if reason:
                rows.append(
                    f"<div style='margin:4px 0;font-size:12px'>"
                    f"<strong style='color:#555'>{label}:</strong> {h(reason)}</div>"
                )
        return "".join(rows) if rows else "<p style='color:#888;font-style:italic'>No reasons recorded</p>"

    # Build table rows
    table_rows = []
    for c in cases:
        tid    = h(c["id"])
        q      = h(c["question"])
        passed = c.get("overall_passed", False)
        row_bg = "#eafaf1" if passed else "#fdf0f0"
        badge_style = (
            "background:#27ae60;color:#fff;padding:2px 8px;border-radius:10px;font-size:11px"
            if passed else
            "background:#e74c3c;color:#fff;padding:2px 8px;border-radius:10px;font-size:11px"
        )
        badge = "PASS" if passed else "FAIL"

        fm  = _metric(c, "Faithfulness")
        arm = _metric(c, "AnswerRelevancy")
        cpm = _metric(c, "ContextualPrecision")
        crm = _metric(c, "ContextualRecall")

        def score_cell(m):
            s = m.get("score", 0.0)
            p = m.get("passed", False)
            col = "#27ae60" if p else "#e74c3c"
            return f"<td style='text-align:center;color:{col};font-weight:bold'>{s:.2f}</td>"

        expand_id = f"detail_{tid}"
        detail_html = (
            f"<tr id='{expand_id}' style='display:none'>"
            f"<td colspan='7' style='background:{row_bg};padding:12px 20px;border-bottom:2px solid #ddd'>"
            f"<div style='display:grid;grid-template-columns:1fr 1fr;gap:16px'>"

            f"<div>"
            f"<p style='font-weight:bold;color:#2c3e50;margin:0 0 4px'>Expected Answer</p>"
            f"<div style='background:#fff;border:1px solid #ddd;padding:8px;border-radius:4px;"
            f"font-size:13px;max-height:120px;overflow-y:auto'>{h(c.get('expected_output',''))}</div>"
            f"</div>"

            f"<div>"
            f"<p style='font-weight:bold;color:#2c3e50;margin:0 0 4px'>Agent Answer</p>"
            f"<div style='background:#fff;border:1px solid #ddd;padding:8px;border-radius:4px;"
            f"font-size:13px;max-height:120px;overflow-y:auto'>{h(c.get('actual_output',''))}</div>"
            f"</div>"

            f"<div style='grid-column:span 2'>"
            f"<p style='font-weight:bold;color:#2c3e50;margin:0 0 4px'>"
            f"Retrieved Chunks ({len(c.get('retrieved_chunks',[]))})</p>"
            f"{_chunk_rows(c.get('retrieved_chunks', []))}"
            f"</div>"

            f"<div style='grid-column:span 2'>"
            f"<p style='font-weight:bold;color:#2c3e50;margin:0 0 4px'>Metric Reasons</p>"
            f"{_reason_rows(c.get('metric_reasons', {}))}"
            f"</div>"

            f"</div>"
            f"</td></tr>"
        )

        row = (
            f"<tr style='background:{row_bg};cursor:pointer' "
            f"onclick=\"toggleRow('{expand_id}')\">"
            f"<td style='padding:8px 12px;font-weight:bold;white-space:nowrap'>{tid}</td>"
            f"<td style='padding:8px 12px;max-width:380px'>{q}</td>"
            f"{score_cell(fm)}{score_cell(arm)}{score_cell(cpm)}{score_cell(crm)}"
            f"<td style='text-align:center;padding:8px 12px'>"
            f"<span style='{badge_style}'>{badge}</span></td>"
            f"</tr>"
            f"{detail_html}"
        )
        table_rows.append(row)

    def summary_card(label, value, color="#2c3e50"):
        return (
            f"<div style='background:#fff;border-radius:8px;padding:16px 24px;"
            f"box-shadow:0 2px 8px rgba(0,0,0,0.08);text-align:center;min-width:130px'>"
            f"<div style='font-size:28px;font-weight:bold;color:{color}'>{value}</div>"
            f"<div style='font-size:13px;color:#666;margin-top:4px'>{label}</div>"
            f"</div>"
        )

    pass_color = "#27ae60" if pass_pct >= 80 else "#e67e22" if pass_pct >= 60 else "#e74c3c"

    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>RAG Eval Report — {run_date}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
         background: #f0f2f5; color: #2c3e50; }}
  .header {{ background: linear-gradient(135deg, #1a2980, #26d0ce);
             color: white; padding: 28px 40px; }}
  .header h1 {{ font-size: 22px; font-weight: 600; }}
  .header .meta {{ font-size: 13px; opacity: 0.85; margin-top: 6px; }}
  .container {{ max-width: 1200px; margin: 0 auto; padding: 24px 20px; }}
  .cards {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 24px; }}
  table {{ width: 100%; border-collapse: collapse; background: white;
           border-radius: 8px; overflow: hidden;
           box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
  thead tr {{ background: #1a2980; color: white; }}
  thead th {{ padding: 12px 14px; text-align: center; font-size: 13px;
              font-weight: 600; white-space: nowrap; }}
  thead th:nth-child(2) {{ text-align: left; }}
  tbody tr:hover {{ filter: brightness(0.97); }}
  tbody td {{ padding: 10px 12px; border-bottom: 1px solid #eee;
              font-size: 13px; vertical-align: top; }}
  .footer {{ text-align: center; font-size: 12px; color: #aaa; margin-top: 24px; padding-bottom: 32px; }}
</style>
</head>
<body>

<div class="header">
  <h1>SG Property Advisory AI — RAG Evaluation Report</h1>
  <div class="meta">Run date: {run_date} &nbsp;|&nbsp; Backend: {h(_BACKEND_URL)} &nbsp;|&nbsp; Pinecone index: {h(active_index)} &nbsp;|&nbsp; Threshold: {_THRESHOLD}</div>
</div>

<div class="container">
  <div class="cards">
    {summary_card("Total Questions", total)}
    {summary_card("Passed", passed, "#27ae60")}
    {summary_card("Pass Rate", f"{pass_pct}%", pass_color)}
    {summary_card("Faithfulness", f"{f_mean:.2f}", score_color(f_mean))}
    {summary_card("Answer Relevancy", f"{ar_mean:.2f}", score_color(ar_mean))}
    {summary_card("Ctx Precision", f"{cp_mean:.2f}", score_color(cp_mean))}
    {summary_card("Ctx Recall", f"{cr_mean:.2f}", score_color(cr_mean))}
  </div>

  <p style="font-size:13px;color:#888;margin-bottom:12px">
    Click any row to expand details (expected answer, agent answer, retrieved chunks, metric reasons).
  </p>

  <table>
    <thead>
      <tr>
        <th>ID</th>
        <th style="text-align:left">Question</th>
        <th>Faithfulness</th>
        <th>Ans Relevancy</th>
        <th>Ctx Precision</th>
        <th>Ctx Recall</th>
        <th>Result</th>
      </tr>
    </thead>
    <tbody>
      {"".join(table_rows)}
    </tbody>
  </table>

  <div class="footer">
    Generated by run_rag_eval.py &nbsp;|&nbsp; {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
  </div>
</div>

<script>
function toggleRow(id) {{
  var el = document.getElementById(id);
  if (el) el.style.display = (el.style.display === 'none' || el.style.display === '') ? 'table-row' : 'none';
}}
</script>
</body>
</html>"""

    report_path.write_text(html_content, encoding="utf-8")
    return report_path


# ── Excel report ──────────────────────────────────────────────────────────────

_GREEN = "FF92D050"
_RED   = "FFFF0000"
_GREY  = "FFD3D3D3"


def _save_excel(cases: list[dict]) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  openpyxl not installed — skipping Excel output. Run: pip install openpyxl")
        return

    if _EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(_EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if _SHEET_NAME in wb.sheetnames:
        del wb[_SHEET_NAME]
    ws = wb.create_sheet(_SHEET_NAME)

    headers = [
        "Test ID", "Question", "Expected Output", "Actual Output",
        "Faithfulness", "F-Pass",
        "AnswerRelevancy", "AR-Pass",
        "CtxPrecision", "CP-Pass",
        "CtxRecall", "CR-Pass",
        "Overall", "Latency (s)", "Date",
    ]

    bold_font    = Font(bold=True)
    header_fill  = PatternFill("solid", fgColor="FF4472C4")
    header_font  = Font(bold=True, color="FFFFFFFF")
    green_fill   = PatternFill("solid", fgColor=_GREEN)
    red_fill     = PatternFill("solid", fgColor=_RED)
    grey_fill    = PatternFill("solid", fgColor=_GREY)
    centre_align = Alignment(horizontal="center")
    wrap_align   = Alignment(wrap_text=True, vertical="top")

    for col_idx, header in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = centre_align

    run_date = date.today().isoformat()

    metric_class_keys = [
        "FaithfulnessMetric",
        "AnswerRelevancyMetric",
        "ContextualPrecisionMetric",
        "ContextualRecallMetric",
    ]

    for row_idx, c in enumerate(cases, start=2):
        values = [
            c.get("id", ""),
            c.get("question", ""),
            c.get("expected_output", ""),
            c.get("actual_output", ""),
        ]

        overall_passed = c.get("overall_passed", False)

        for mk in metric_class_keys:
            m = c.get(mk, {})
            values.append(round(m.get("score", 0.0), 2))
            values.append("PASS" if m.get("passed") else "FAIL")

        values += [
            "PASS" if overall_passed else "FAIL",
            c.get("latency_s", 0.0),
            run_date,
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx in (2, 3, 4):
                cell.alignment = wrap_align
            elif col_idx in (6, 8, 10, 12):
                cell.fill      = green_fill if val == "PASS" else red_fill
                cell.alignment = centre_align
            elif col_idx == 13:
                cell.fill      = green_fill if overall_passed else red_fill
                cell.alignment = centre_align
            else:
                cell.alignment = centre_align

    summary_row = len(cases) + 2
    total  = len(cases)
    passed = sum(1 for c in cases if c.get("overall_passed"))

    ws.cell(row=summary_row, column=1, value="SUMMARY").font = bold_font
    ws.cell(row=summary_row, column=13, value=f"{passed}/{total} passed").font = bold_font

    for col_idx, mk in zip([5, 7, 9, 11], metric_class_keys):
        scores = [c.get(mk, {}).get("score", 0.0) for c in cases]
        mean   = round(sum(scores) / len(scores), 2) if scores else 0.0
        cell   = ws.cell(row=summary_row, column=col_idx, value=mean)
        cell.font      = bold_font
        cell.fill      = grey_fill
        cell.alignment = centre_align

    col_widths = [10, 50, 40, 50, 14, 8, 16, 8, 14, 8, 12, 8, 10, 12, 12]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.freeze_panes = "B2"
    wb.save(_EXCEL_FILE)


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    records = _load_testset()
    total   = len(records)
    print(f"Loaded {total} test case(s) from {_TESTSET_FILE.name}")
    print(f"  Backend          : {_BACKEND_URL}")
    active_index = _check_kb_environment()
    print()

    cases = []
    for i, rec in enumerate(records):
        test_id = f"TC-{i+1:03d}"
        if not rec["input"].strip():
            print(f"  {test_id}  SKIP (empty question)")
            continue
        cases.append(_evaluate_case(test_id, rec))

    passed    = sum(1 for c in cases if c.get("overall_passed"))
    pass_rate = round(passed / len(cases) * 100, 1) if cases else 0.0

    print(f"\n{'═'*60}")
    print(f"  Total:       {len(cases)}")
    print(f"  Passed:      {passed}")
    print(f"  Pass rate:   {pass_rate}%")

    run_date = date.today().isoformat()

    json_path = _save_json(cases)
    print(f"\nJSON report  → {json_path}")

    html_path = _save_html(cases, run_date, active_index)
    print(f"HTML report  → {html_path}  (open in any browser)")

    _save_excel(cases)
    print(f'Excel report → {_EXCEL_FILE}  (sheet: "{_SHEET_NAME}")')


if __name__ == "__main__":
    main()
